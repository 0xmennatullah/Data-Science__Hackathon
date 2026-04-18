import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
INPUT_FILE  = 'Employees.xlsx'   # change path if needed
OUTPUT_FILE = 'Employees_Cleaned.xlsx'
REF_DATE    = pd.Timestamp('2021-01-01')  # reference date for tenure calc

# ── LOAD ──────────────────────────────────────────────────────────────────────
df = pd.read_excel(INPUT_FILE)
print(f"Loaded {len(df)} rows, {len(df.columns)} columns")

# ── CLEAN ─────────────────────────────────────────────────────────────────────
df['First Name'] = df['First Name'].str.strip().str.title()
df['Last Name']  = df['Last Name'].str.strip().str.title()
df['Full Name']  = df['First Name'] + ' ' + df['Last Name']
df['Start Date'] = pd.to_datetime(df['Start Date'])

# Recalculate tenure from Start Date (fixes ~30 borderline records)
df['Tenure Years'] = ((REF_DATE - df['Start Date']).dt.days / 365.25).round().astype(int)

# Salary band labels
bins   = [0, 1000, 1500, 2000, 2500, 9999]
labels = ['<1K', '1K-1.5K', '1.5K-2K', '2K-2.5K', '>2.5K']
df['Salary Band'] = pd.cut(df['Monthly Salary'], bins=bins, labels=labels).astype(str)

# Job level from job rate
rate_map = {1.0: 'Junior', 2.0: 'Mid', 3.0: 'Senior', 4.5: 'Lead', 5.0: 'Principal'}
df['Job Level'] = df['Job Rate'].map(rate_map)

# Leave & overtime flags
df['Has Sick Leave']   = df['Sick Leaves']   > 0
df['Has Unpaid Leave'] = df['Unpaid Leaves']  > 0
df['High Overtime']    = df['Overtime Hours'] > 50   # flag if > 50 hrs

clean_cols = [
    'No', 'Full Name', 'First Name', 'Last Name', 'Gender', 'Start Date',
    'Tenure Years', 'Department', 'Country', 'Center',
    'Monthly Salary', 'Annual Salary', 'Salary Band',
    'Job Rate', 'Job Level',
    'Sick Leaves', 'Unpaid Leaves', 'Overtime Hours',
    'Has Sick Leave', 'Has Unpaid Leave', 'High Overtime'
]
df_clean = df[clean_cols].copy()
total = len(df_clean)
print(f"Clean dataframe: {df_clean.shape}")

# ── STYLES ────────────────────────────────────────────────────────────────────
NAVY   = PatternFill('solid', start_color='1F3864')
BLUE   = PatternFill('solid', start_color='2E75B6')
LBLUE  = PatternFill('solid', start_color='D6E4F0')
ACCENT = PatternFill('solid', start_color='EBF3FB')
WHITE  = PatternFill('solid', start_color='FFFFFF')
GREEN  = PatternFill('solid', start_color='E2EFDA')
AMBER  = PatternFill('solid', start_color='FFF2CC')
RED    = PatternFill('solid', start_color='FCE4D6')

thin = Side(style='thin', color='B0C4DE')
bord = Border(left=thin, right=thin, top=thin, bottom=thin)

def sh(c, fill=NAVY, size=10, bold=True, color='FFFFFF'):
    c.fill = fill
    c.font = Font(name='Arial', bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = bord

def sc(c, fill=WHITE, bold=False, align='left'):
    c.fill = fill
    c.font = Font(name='Arial', bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = bord

# ── WORKBOOK ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══ SHEET 1: Clean Data ══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Clean Data'

for ci, h in enumerate(df_clean.columns, 1):
    c = ws1.cell(1, ci, h)
    sh(c)
ws1.row_dimensions[1].height = 28

for ri, row in enumerate(df_clean.itertuples(index=False), 2):
    for ci, val in enumerate(row, 1):
        if isinstance(val, bool):
            val = 'Yes' if val else 'No'
        c = ws1.cell(ri, ci, val)
        sc(c, fill=ACCENT if ri % 2 == 0 else WHITE)

col_widths = [5, 22, 12, 12, 8, 14, 10, 22, 16, 8, 14, 14, 10, 8, 10, 10, 12, 14, 12, 14, 12]
for ci, w in enumerate(col_widths[:len(df_clean.columns)], 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w
ws1.freeze_panes = 'A2'

# ═══ SHEET 2: Insights ════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Insights')
ws2.sheet_view.showGridLines = False
for ci in range(1, 16):
    ws2.column_dimensions[get_column_letter(ci)].width = 15

# Title
ws2.merge_cells('A1:O1')
c = ws2['A1']
c.value = 'HR Employee Insights Dashboard'
c.fill = NAVY
c.font = Font(name='Arial', bold=True, size=16, color='FFFFFF')
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = bord
ws2.row_dimensions[1].height = 36

# KPIs
avg_salary  = int(df_clean['Monthly Salary'].mean())
avg_tenure  = round(df_clean['Tenure Years'].mean(), 1)
sick_rate   = round(df_clean['Has Sick Leave'].mean() * 100, 1)
unpaid_rate = round(df_clean['Has Unpaid Leave'].mean() * 100, 1)
high_ot     = round(df_clean['High Overtime'].mean() * 100, 1)

kpis = [
    ('Total Employees',    total,           LBLUE),
    ('Avg Monthly Salary', f'${avg_salary:,}', LBLUE),
    ('Avg Tenure (yrs)',   avg_tenure,      GREEN),
    ('Sick Leave Rate',    f'{sick_rate}%', AMBER),
    ('Unpaid Leave Rate',  f'{unpaid_rate}%', AMBER),
    ('High Overtime %',    f'{high_ot}%',   RED),
    ('Departments',        20,              LBLUE),
    ('Countries',          5,               LBLUE),
]
for i, (label, val, bg) in enumerate(kpis):
    col = i * 2 + 1
    if col > 15:
        break
    ws2.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    c = ws2.cell(3, col, label)
    sh(c, fill=BLUE, size=9)
    ws2.row_dimensions[3].height = 16
    ws2.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col+1)
    c2 = ws2.cell(4, col, val)
    sh(c2, fill=bg, size=14, color='1F3864')
    ws2.row_dimensions[4].height = 24
    ws2.row_dimensions[5].height = 10

# Dept salary (cols 1–7)
dept_sal = df_clean.groupby('Department')['Monthly Salary'].mean().sort_values(ascending=False).round(0)
ws2.merge_cells('A7:G7')
c = ws2['A7']
c.value = 'Avg Monthly Salary by Department'
sh(c, size=12)
ws2.row_dimensions[7].height = 26
sh(ws2.cell(8, 1, 'Department'), fill=BLUE)
ws2.merge_cells('B8:F8')
sh(ws2.cell(8, 2, ''), fill=BLUE)
sh(ws2.cell(8, 7, 'Avg ($)'), fill=BLUE)
r = 9
for dept, sal in dept_sal.items():
    bg = ACCENT if r % 2 == 0 else WHITE
    sc(ws2.cell(r, 1, dept), fill=bg)
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    sc(ws2.cell(r, 2, ''), fill=bg)
    sc(ws2.cell(r, 7, int(sal)), fill=bg, align='center')
    r += 1

# Gender/Country (cols 9–15)
gender_dist  = df_clean['Gender'].value_counts()
country_dist = df_clean['Country'].value_counts()
ws2.merge_cells('I7:O7')
c = ws2['I7']
c.value = 'Gender & Country Distribution'
sh(c, size=12)
sh(ws2.cell(8, 9,  'Category'), fill=BLUE)
ws2.merge_cells('J8:M8')
sh(ws2.cell(8, 10, 'Group'),    fill=BLUE)
sh(ws2.cell(8, 14, 'Count'),    fill=BLUE)
sh(ws2.cell(8, 15, '%'),        fill=BLUE)
r2 = 9
for g, cnt in gender_dist.items():
    bg = ACCENT if r2 % 2 == 0 else WHITE
    sc(ws2.cell(r2, 9,  'Gender'), fill=bg, align='center')
    ws2.merge_cells(start_row=r2, start_column=10, end_row=r2, end_column=13)
    sc(ws2.cell(r2, 10, g),   fill=bg, align='center')
    sc(ws2.cell(r2, 14, cnt), fill=bg, align='center')
    sc(ws2.cell(r2, 15, f'{cnt/total*100:.1f}%'), fill=bg, align='center')
    r2 += 1
for cntry, cnt in country_dist.items():
    bg = ACCENT if r2 % 2 == 0 else WHITE
    sc(ws2.cell(r2, 9,  'Country'), fill=bg, align='center')
    ws2.merge_cells(start_row=r2, start_column=10, end_row=r2, end_column=13)
    sc(ws2.cell(r2, 10, cntry), fill=bg, align='center')
    sc(ws2.cell(r2, 14, cnt),   fill=bg, align='center')
    sc(ws2.cell(r2, 15, f'{cnt/total*100:.1f}%'), fill=bg, align='center')
    r2 += 1

# Overtime by dept (cols 9–15)
r2 += 1
ws2.merge_cells(start_row=r2, start_column=9, end_row=r2, end_column=15)
c = ws2.cell(r2, 9, 'Avg Overtime Hours by Department')
sh(c, size=12)
r2 += 1
sh(ws2.cell(r2, 9, 'Department'), fill=BLUE)
ws2.merge_cells(start_row=r2, start_column=10, end_row=r2, end_column=14)
sh(ws2.cell(r2, 10, ''), fill=BLUE)
sh(ws2.cell(r2, 15, 'Avg OT Hrs'), fill=BLUE)
r2 += 1
ot_dept = df_clean.groupby('Department')['Overtime Hours'].mean().sort_values(ascending=False).round(1)
for dept, ot in ot_dept.head(10).items():
    bg = RED if ot > 20 else (AMBER if ot > 10 else WHITE)
    sc(ws2.cell(r2, 9, dept), fill=bg)
    ws2.merge_cells(start_row=r2, start_column=10, end_row=r2, end_column=14)
    sc(ws2.cell(r2, 10, ''), fill=bg)
    sc(ws2.cell(r2, 15, float(ot)), fill=bg, align='center')
    r2 += 1

# Job Level + Salary Band (below dept salary)
r3 = r + 2
ws2.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=7)
c = ws2.cell(r3, 1, 'Job Level & Salary Band')
sh(c, size=12)
r3 += 1
sh(ws2.cell(r3, 1, 'Job Level'), fill=BLUE)
ws2.merge_cells(start_row=r3, start_column=2, end_row=r3, end_column=4)
sh(ws2.cell(r3, 2, 'Count'), fill=BLUE)
sh(ws2.cell(r3, 5, 'Salary Band'), fill=BLUE)
ws2.merge_cells(start_row=r3, start_column=6, end_row=r3, end_column=7)
sh(ws2.cell(r3, 6, 'Count'), fill=BLUE)
r3 += 1
level_dist  = df_clean['Job Level'].value_counts()
salary_band = df_clean['Salary Band'].value_counts()
lv = list(level_dist.items())
sb = list(salary_band.items())
for i in range(max(len(lv), len(sb))):
    bg = ACCENT if r3 % 2 == 0 else WHITE
    if i < len(lv):
        sc(ws2.cell(r3, 1, lv[i][0]), fill=bg)
        ws2.merge_cells(start_row=r3, start_column=2, end_row=r3, end_column=4)
        sc(ws2.cell(r3, 2, lv[i][1]), fill=bg, align='center')
    if i < len(sb):
        sc(ws2.cell(r3, 5, sb[i][0]), fill=bg)
        ws2.merge_cells(start_row=r3, start_column=6, end_row=r3, end_column=7)
        sc(ws2.cell(r3, 6, sb[i][1]), fill=bg, align='center')
    r3 += 1

# Gender salary gap
r3 += 1
ws2.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=7)
c = ws2.cell(r3, 1, 'Gender Salary Comparison')
sh(c, size=12)
r3 += 1
sh(ws2.cell(r3, 1, 'Gender'), fill=BLUE)
ws2.merge_cells(start_row=r3, start_column=2, end_row=r3, end_column=4)
sh(ws2.cell(r3, 2, 'Avg Monthly Salary'), fill=BLUE)
ws2.merge_cells(start_row=r3, start_column=5, end_row=r3, end_column=7)
sh(ws2.cell(r3, 5, 'Avg Annual Salary'), fill=BLUE)
r3 += 1
gender_sal = df_clean.groupby('Gender')['Monthly Salary'].mean().round(0)
for g, s in gender_sal.items():
    bg = ACCENT if r3 % 2 == 0 else WHITE
    sc(ws2.cell(r3, 1, g), fill=bg, align='center')
    ws2.merge_cells(start_row=r3, start_column=2, end_row=r3, end_column=4)
    sc(ws2.cell(r3, 2, f'${int(s):,}'), fill=bg, align='center')
    ws2.merge_cells(start_row=r3, start_column=5, end_row=r3, end_column=7)
    sc(ws2.cell(r3, 5, f'${int(s*12):,}'), fill=bg, align='center')
    r3 += 1

# ═══ SHEET 3: Summary Tables (Power BI ready) ════════════════════════════════
ws3 = wb.create_sheet('Summary Tables')
ws3.sheet_view.showGridLines = False

def write_flat_table(ws, start_r, start_c, title, data_df):
    ncols = len(data_df.columns)
    ws.merge_cells(start_row=start_r, start_column=start_c,
                   end_row=start_r, end_column=start_c+ncols-1)
    c = ws.cell(start_r, start_c, title)
    sh(c, size=12)
    ws.row_dimensions[start_r].height = 24
    for ci, col in enumerate(data_df.columns, start_c):
        c = ws.cell(start_r+1, ci, col)
        sh(c, fill=BLUE)
        ws.column_dimensions[get_column_letter(ci)].width = 20
    for ri, row in enumerate(data_df.itertuples(index=False), start_r+2):
        for ci, val in enumerate(row, start_c):
            c = ws.cell(ri, ci, val)
            sc(c, fill=ACCENT if ri % 2 == 0 else WHITE, align='center')
    return start_r + len(data_df) + 4

dept_sum = df_clean.groupby('Department').agg(
    Employees        =('No', 'count'),
    Avg_Monthly_Salary=('Monthly Salary', lambda x: round(x.mean(), 0)),
    Total_Annual_Cost =('Annual Salary', 'sum'),
    Avg_OT_Hours      =('Overtime Hours', lambda x: round(x.mean(), 1)),
    Sick_Leave_Count  =('Has Sick Leave', 'sum'),
    High_OT_Count     =('High Overtime', 'sum'),
).reset_index()
dept_sum.columns = ['Department', 'Employees', 'Avg Monthly Salary ($)',
                    'Total Annual Cost ($)', 'Avg OT Hours', 'Sick Leave #', 'High OT #']

country_sum = df_clean.groupby('Country').agg(
    Employees  =('No', 'count'),
    Avg_Salary =('Monthly Salary', lambda x: round(x.mean(), 0)),
    Avg_OT     =('Overtime Hours', lambda x: round(x.mean(), 1)),
).reset_index()
country_sum.columns = ['Country', 'Employees', 'Avg Monthly Salary ($)', 'Avg OT Hours']

gl_sum = df_clean.groupby(['Gender', 'Job Level']).agg(
    Count     =('No', 'count'),
    Avg_Salary=('Monthly Salary', lambda x: round(x.mean(), 0)),
).reset_index()
gl_sum.columns = ['Gender', 'Job Level', 'Count', 'Avg Monthly Salary ($)']

center_sum = df_clean.groupby('Center').agg(
    Employees =('No', 'count'),
    Avg_Salary=('Monthly Salary', lambda x: round(x.mean(), 0)),
    Avg_OT    =('Overtime Hours', lambda x: round(x.mean(), 1)),
).reset_index()
center_sum.columns = ['Center', 'Employees', 'Avg Monthly Salary ($)', 'Avg OT Hours']

r = 2
r = write_flat_table(ws3, r, 1, 'Department Summary', dept_sum)
r = write_flat_table(ws3, r, 1, 'Country Summary', country_sum)
r = write_flat_table(ws3, r, 1, 'Gender x Job Level', gl_sum)
r = write_flat_table(ws3, r, 1, 'Center Summary', center_sum)

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"Saved → {OUTPUT_FILE}")